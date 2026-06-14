import streamlit as st
import pandas as pd
import json
import uuid

from database import *
from email_service import *
from openpyxl import load_workbook

# -------------------------
# PAGE CONFIG
# -------------------------

st.set_page_config(
    page_title="InformAI Forms",
    layout="wide"
)

init_db()
# ==========================================
# PUBLIC FORM PAGE
# ==========================================

query_params = st.query_params

if "form_id" in query_params:

    form_id = query_params["form_id"]

    form = get_form(form_id)

    if form:

        st.title(form["form_name"])

        fields = json.loads(
            form["columns_json"]
        )

        response_data = {}

        for field in fields:

            field_name = field["name"]

            values = field.get(
                "values",
                []
            )

            if len(values) > 1:

                response_data[field_name] = st.selectbox(
                    field_name,
                    values,
                    key=field_name
                )

            else:

                response_data[field_name] = st.text_input(
                    field_name,
                    key=field_name
                )

        if st.button("Submit Form"):

            save_response(
                str(uuid.uuid4()),
                form_id,
                json.dumps(response_data)
            )

            st.success(
                "Form Submitted Successfully"
            )

        st.stop()

    else:

        st.error("Form Not Found")
        st.stop()

# -------------------------
# SESSION STATE
# -------------------------

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if "user" not in st.session_state:
    st.session_state.user = None

# -------------------------
# TITLE
# -------------------------

st.title("📄 InformAI Forms")
st.caption("Excel → Dynamic Form → Email → Responses")

# =====================================================
# LOGIN / REGISTER
# =====================================================

if not st.session_state.logged_in:

    tab1, tab2 = st.tabs(["Login", "Register"])

    # ---------------- LOGIN ----------------

    with tab1:

        st.subheader("Login")

        login_email = st.text_input(
            "Email",
            key="login_email"
        )

        login_password = st.text_input(
            "Password",
            type="password",
            key="login_password"
        )

        if st.button("Login"):

            user = login_user(
                login_email,
                login_password
            )

            if user:

                st.session_state.logged_in = True
                st.session_state.user = dict(user)

                st.success("Login Successful")
                st.rerun()

            else:
                st.error("Invalid Credentials")

    # ---------------- REGISTER ----------------

    with tab2:

        st.subheader("Register")

        username = st.text_input("Username")
        email = st.text_input("Email")
        password = st.text_input(
            "Password",
            type="password"
        )

        if st.button("Create Account"):

            result = create_user(
                username,
                email,
                password
            )

            if result:
                st.success(
                    "Account Created Successfully"
                )
            else:
                st.error(
                    "Email already exists"
                )

    st.stop()

# =====================================================
# DASHBOARD
# =====================================================

user = st.session_state.user

st.sidebar.success(
    f"Welcome {user['username']}"
)

if st.sidebar.button("Logout"):

    st.session_state.logged_in = False
    st.session_state.user = None

    st.rerun()

st.header("Dashboard")

form_name = st.text_input(
    "Form Name"
)

uploaded_file = st.file_uploader(
    "Upload Excel File",
    type=["xlsx"]
)

if uploaded_file:

    wb = load_workbook(uploaded_file)

    ws = wb.active

    # Header detection from first 5 rows
    header_row = None

    for row_num in range(1, 6):

        values = [
            cell.value
            for cell in ws[row_num]
            if cell.value is not None
        ]

        if len(values) >= 2:

            header_row = row_num
            break

    if header_row is None:
        header_row = 1

    # Reload file pointer
    uploaded_file.seek(0)

    df = pd.read_excel(
        uploaded_file,
        header=header_row - 1
    )

    st.success(
        "Excel Loaded Successfully"
    )

    st.write(
        f"Header Row Found: {header_row}"
    )

    st.subheader(
        "Detected Columns"
    )

    st.dataframe(
        pd.DataFrame(
            {
                "Columns": df.columns
            }
        )
    )

    # ==========================
    # Dropdown Detection
    # ==========================

    dropdown_columns = {}

    try:

        for dv in ws.data_validations.dataValidation:

            if dv.type == "list":

                options = dv.formula1

                for cell_range in dv.cells.ranges:

                    col_num = cell_range.min_col

                    col_name = ws.cell(
                        header_row,
                        col_num
                    ).value

                    dropdown_columns[
                        str(col_name)
                    ] = options

    except:
        pass

    st.subheader(
        "Detected Dropdowns"
    )

    if dropdown_columns:

        for col, values in dropdown_columns.items():

            st.write(
                f"{col} → {values}"
            )

    else:

        st.info(
            "No dropdown lists found."
        )

    # ==========================
    # Save Form Data
    # ==========================

    columns_data = []

    for col in df.columns:

        values = (
            df[col]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        columns_data.append(
            {
                "name": col,
                "values": values[:20]
            }
        )

    if st.button("Save Form"):

        form_id = str(uuid.uuid4())

        save_form(
            form_id=form_id,
            user_id=user["id"],
            form_name=form_name,
            columns_json=json.dumps(
                columns_data
            )
        )

        st.success(
            f"Form Created Successfully: {form_id}"
        )
    
# =====================================================
# MY FORMS
# =====================================================
# =====================================================
# MY FORMS
# =====================================================

st.subheader("My Forms")

forms = get_user_forms(user["id"])

if forms:

    for form in forms:

        st.subheader(form["form_name"])

        if st.button(
            "Generate Link",
            key=f"link_{form['id']}"
        ):

            form_link = (
                form_link = f"?form_id={form['id']}"
                st.code(form_link)
            )

            st.success("Form Link Generated")

            st.code(form_link)

else:

    st.warning(
        "No forms created yet."
    )

   
